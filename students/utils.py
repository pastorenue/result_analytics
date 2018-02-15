from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from result_analytics.utils.excel import ExcelReport
from .models import Student, UniqueMapper
from django.db import transaction
from courses.models import *
from core.models import Activation, StudentSetup
from students.models import Student
from results.models import Result
import uuid


#Defined Functions for students

def create_user(first_name, last_name):
    """Creates a user with a username generated from the supplied `first_name` and `last_name`."""

    user = None
    user = User.objects.create(first_name=first_name, last_name=last_name)
    return user



def export_performance_excel():
    all_cgpa = main.get_performance_report()

    students = Student.objects.all()
    print(all_cgpa)
    fcgpa = ((student.full_name, student.reg_number, all_cgpa[student]) for student in students)
    fields = ["student","reg number","FCGPA"]
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment;filename=performance_report.xls"
    report = ExcelReport(fcgpa, fields, groupby=request.GET.get('groupby'))
    report.write(response)
    return response

def export_transcript_pdf():
    pass


def user_is_student(user):
    return Student.objects.filter(user=user).exists()


@transaction.atomic
def generate_mapper_excel(excel_file):
    from xlrd import open_workbook
    from openpyxl import load_workbook
    from xlutils.copy import copy
    import xlwt
    import urllib

    read_book = open_workbook(file_contents=excel_file.read())
    work_book = copy(read_book)
    sheet = read_book.sheet_by_index(0)
    new_sheet = work_book.get_sheet(0)
    headers = ['REG NUMBER', 'SHORT INSTITUTION CODE', 'REG MAPPER']
    mapper_column = []
    style1 = xlwt.easyxf('pattern: pattern solid, fore_colour red;')

    for row in range(1, sheet.nrows):
        reg_number = sheet.cell(row, 0).value
        short_institution_name = sheet.cell(row, 1).value
        mapper = UniqueMapper(reg_number=reg_number, short_institution_name=short_institution_name)
        mapper.save()
        mapper_column.append(mapper.unique_map)

    for row in range(len(mapper_column)):
        if row == len(mapper_column):
            break
        else:
            new_sheet.write(row+1, 2, mapper_column[row])
    for col in range(len(headers)):
        new_sheet.write(0,col, headers[col], style1)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "inline; filename=%s.xls" % (excel_file.name[:-3])

    work_book.save(response)
    return response

def has_completed_level(student):
    results = Result.objects.filter(student=student, level=student.level)
    reg_courses = CourseRegistration.objects.filter(student=student, level=student.level)

    set_result = set(results)
    set_registration = set(reg_courses)

    set_differences = set_registration.difference(set_result)
    if len(set_differences) > 0:
        return False
    return False


@transaction.atomic
def import_student_from_csv(csv_file, lecturer):
    """
    Import student CSV data.

    """
    csv_data = []
    ifile = csv_file.read().decode("utf-8")
    for row in ifile.split('\n'):
        csv_data.append(row.split(','))

    result_objects = []
    # Check if headers exists. Skip the first entry if true.
    header_check = ['reg_number', 'surname', 'first_name', 'sex', 'level', 'department', 'year_of_admission', 'course_duration']
    first_row = [i.lower().strip() for i in csv_data[0]]
    # if all(i in first_row for i in header_check):
    csv_data = csv_data[1:]

    new_value = 0
    existing_value = 0 # To get the number of records entered
    for row in csv_data:
        # Let's do an extra check to make sure the row is not empty.
        if row:
            department = Department.objects.get(code=str(row[5]).upper())
            if Student.objects.filter(reg_number=row[0]).exists():
                existing_value+=1
            else:
                user = create_user(row[1], row[2])
                user.username = row[0]
                user.password = row[0]
                user.save()

                student = Student(
                    reg_number =row[0],
                    user=user,
                    last_name = row[1],
                    first_name = row[2],
                    department=department,
                    institution=lecturer.institution,
                    year_of_admission=row[6],
                    level=row[4],
                    course_duration=row[7],
                    sex=row[3],
                )
                setup = StudentSetup(user=user)
                setup.save()

                activation = Activation(user=user)
                activation.save()
                create_slug(student)
                new_value+=1
    return new_value, existing_value


def create_slug(instance):
    #import slugify
    from django.utils.text import slugify

    orig = slugify(instance.last_name)
    if Student.objects.filter(slug=instance.slug).exists():
        instance.slug = "%s-%s" % (orig, uuid.uuid4())
    else:
        instance.slug = "%s-%s" % (orig, uuid.uuid4())

    instance.save()
